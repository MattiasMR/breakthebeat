"""Send one privacy-safe BCC request to active Break The Beat participants.

Dry-run is the default. Add --send only after reviewing the workbook and count.
"""

from __future__ import annotations

import argparse
import os
import smtplib
import ssl
import subprocess
import sys
import unicodedata
import urllib.parse
import webbrowser
from email.message import EmailMessage
from pathlib import Path

from openpyxl import load_workbook


ACTIVE_STATUSES = {"activo", "active", "confirmado", "confirmed"}
PHOTO_COMPLETE = {"cargada", "cargado", "si", "sí", "yes", "completa", "completada"}


def normalized(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").strip().lower())
    return "".join(character for character in text if not unicodedata.combining(character))


def active_recipient_emails(workbook_path: Path, only_missing_photo: bool = True) -> list[str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook["Participantes"] if "Participantes" in workbook.sheetnames else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration as error:
            raise ValueError("El Excel no contiene filas.") from error

        columns = {normalized(value): index for index, value in enumerate(headers)}
        if "email" not in columns or "estado" not in columns:
            raise ValueError("El Excel debe contener las columnas Email y Estado.")
        photo_index = columns.get("foto")
        recipients: dict[str, None] = {}
        for row in rows:
            status = normalized(row[columns["estado"]] if columns["estado"] < len(row) else "")
            if status not in ACTIVE_STATUSES:
                continue
            if only_missing_photo and photo_index is not None:
                photo_status = normalized(row[photo_index] if photo_index < len(row) else "")
                if photo_status in {normalized(value) for value in PHOTO_COMPLETE}:
                    continue
            email = str(row[columns["email"]] or "").strip().lower()
            if "@" in email and "." in email.rsplit("@", 1)[-1]:
                recipients[email] = None
        return list(recipients)
    finally:
        workbook.close()


def build_message(sender: str, to_address: str, reply_to: str, upload_url: str) -> EmailMessage:
    message = EmailMessage()
    message["Subject"] = "Importante: completa tu fotografía para Break The Beat 2026"
    if sender:
        message["From"] = sender
    if to_address:
        message["To"] = to_address
    if reply_to:
        message["Reply-To"] = reply_to
    message.set_content(
        f"""Hola,

Para completar tu inscripción en Break The Beat 2026 necesitamos que cargues una fotografía reciente donde se vea claramente tu rostro.

Es importante que lo hagas antes del evento. Ingresa aquí:
{upload_url}

Puedes identificarte con el correo que usaste al inscribirte, con tu código de participante o con el contenido de tu QR. Si participas en varias categorías con el mismo correo, una sola carga actualizará tus registros activos.

Tu fotografía se guardará de forma privada y se utilizará para la organización e identificación de participantes.

Gracias,
Equipo Break The Beat
"""
    )
    message.add_alternative(
        f"""<!doctype html><html lang="es"><body style="font-family:Arial,sans-serif;color:#171717;line-height:1.6">
<h2 style="color:#ff0076">Completa tu fotografía</h2>
<p>Para completar tu inscripción en <strong>Break The Beat 2026</strong> necesitamos una fotografía reciente donde se vea claramente tu rostro.</p>
<p><strong>Es importante que lo hagas antes del evento.</strong></p>
<p><a href="{upload_url}" style="display:inline-block;padding:12px 20px;background:#ff0076;color:white;text-decoration:none;border-radius:6px">Subir mi fotografía</a></p>
<p>Puedes identificarte con el correo de tu inscripción, con tu código de participante o con el contenido de tu QR. Si participas en varias categorías con el mismo correo, una sola carga actualizará tus registros activos.</p>
<p>Tu fotografía se guardará de forma privada y se utilizará para la organización e identificación de participantes.</p>
<p>Gracias,<br>Equipo Break The Beat</p>
</body></html>""",
        subtype="html",
    )
    return message


def plain_message_body(message: EmailMessage) -> str:
    body = message.get_body(preferencelist=("plain",))
    if body is None:
        raise ValueError("No se pudo preparar el cuerpo del correo.")
    return body.get_content().strip()


def gmail_compose_url(message: EmailMessage, to_address: str) -> str:
    query = urllib.parse.urlencode(
        {
            "view": "cm",
            "fs": "1",
            "to": to_address,
            "su": str(message["Subject"] or ""),
            "body": plain_message_body(message),
        }
    )
    return f"https://mail.google.com/mail/?{query}"


def default_mail_app_url(message: EmailMessage, to_address: str) -> str:
    query = urllib.parse.urlencode(
        {"subject": str(message["Subject"] or ""), "body": plain_message_body(message)}
    )
    return f"mailto:{urllib.parse.quote(to_address)}?{query}"


def copy_bcc_to_clipboard(recipients: list[str]) -> None:
    bcc_text = ", ".join(recipients)
    if os.name != "nt":
        raise ValueError("La copia automática al portapapeles está preparada para Windows.")
    subprocess.run(["clip.exe"], input=bcc_text, text=True, check=True)


def open_manual_compose(message: EmailMessage, recipients: list[str], to_address: str, mode: str) -> None:
    copy_bcc_to_clipboard(recipients)
    url = gmail_compose_url(message, to_address) if mode == "gmail" else default_mail_app_url(message, to_address)
    if not webbrowser.open(url, new=2):
        raise ValueError("No se pudo abrir la ventana de correo.")


def send_message(message: EmailMessage, recipients: list[str], args: argparse.Namespace) -> None:
    password = os.getenv("BTB_SMTP_PASSWORD", "")
    username = os.getenv("BTB_SMTP_USER", "")
    if not args.smtp_host or not username or not password:
        raise ValueError("Faltan BTB_SMTP_HOST, BTB_SMTP_USER o BTB_SMTP_PASSWORD.")

    all_recipients = [args.to_address, *recipients]
    if args.smtp_security == "ssl":
        with smtplib.SMTP_SSL(args.smtp_host, args.smtp_port, context=ssl.create_default_context()) as smtp:
            smtp.login(username, password)
            smtp.send_message(message, from_addr=args.sender, to_addrs=all_recipients)
    else:
        with smtplib.SMTP(args.smtp_host, args.smtp_port, timeout=30) as smtp:
            smtp.ehlo()
            smtp.starttls(context=ssl.create_default_context())
            smtp.ehlo()
            smtp.login(username, password)
            smtp.send_message(message, from_addr=args.sender, to_addrs=all_recipients)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Solicita fotografías a participantes activos usando un único correo con CCO.")
    parser.add_argument("excel", type=Path, help="Excel operativo descargado desde el panel de administración.")
    action = parser.add_mutually_exclusive_group()
    action.add_argument("--open-gmail", action="store_true", help="Abre Gmail con el mensaje preparado y copia los destinatarios para pegarlos en CCO.")
    action.add_argument("--open-mail-app", action="store_true", help="Abre la aplicación de correo predeterminada y copia los destinatarios para CCO.")
    action.add_argument("--send", action="store_true", help="Envía directamente por SMTP. Sin una acción solo valida y muestra el total.")
    parser.add_argument("--include-with-photo", action="store_true", help="Incluye también quienes ya aparecen con foto cargada.")
    parser.add_argument("--max-recipients", type=int, default=450, help="Límite de seguridad para un único envío CCO.")
    parser.add_argument("--upload-url", default=os.getenv("BTB_PHOTO_UPLOAD_URL", "https://breakthebeat.com.ec/actualizar-foto/"))
    parser.add_argument("--sender", default=os.getenv("BTB_MAIL_FROM", ""))
    parser.add_argument("--to-address", default=os.getenv("BTB_MAIL_TO", ""), help="Dirección visible en Para; normalmente la del organizador.")
    parser.add_argument("--reply-to", default=os.getenv("BTB_MAIL_REPLY_TO", ""))
    parser.add_argument("--smtp-host", default=os.getenv("BTB_SMTP_HOST", ""))
    parser.add_argument("--smtp-port", type=int, default=int(os.getenv("BTB_SMTP_PORT", "587")))
    parser.add_argument("--smtp-security", choices=("starttls", "ssl"), default=os.getenv("BTB_SMTP_SECURITY", "starttls"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.excel.is_file():
        print(f"No existe el archivo: {args.excel}", file=sys.stderr)
        return 2
    try:
        recipients = active_recipient_emails(args.excel, only_missing_photo=not args.include_with_photo)
        if not recipients:
            print("No hay destinatarios activos que cumplan los filtros.")
            return 0
        if len(recipients) > args.max_recipients:
            raise ValueError(f"Hay {len(recipients)} destinatarios; supera el límite de seguridad de {args.max_recipients}.")
        print(f"Destinatarios únicos en CCO: {len(recipients)}")
        print(f"Página incluida en el correo: {args.upload_url}")
        if not args.send and not args.open_gmail and not args.open_mail_app:
            print("Simulación terminada: no se abrió ni envió ningún correo.")
            print("Usa --open-gmail, --open-mail-app o --send después de revisar el total.")
            return 0
        message = build_message(args.sender, args.to_address, args.reply_to, args.upload_url)
        if args.open_gmail or args.open_mail_app:
            mode = "gmail" if args.open_gmail else "mail"
            open_manual_compose(message, recipients, args.to_address, mode)
            print("Los destinatarios quedaron copiados en el portapapeles.")
            print("Pégalos con Ctrl+V en el campo CCO, revisa el mensaje y pulsa Enviar manualmente.")
            return 0
        if not args.sender or not args.to_address:
            raise ValueError("Faltan BTB_MAIL_FROM o BTB_MAIL_TO.")
        send_message(message, recipients, args)
        print(f"Correo enviado una sola vez con {len(recipients)} destinatarios en CCO.")
        return 0
    except (OSError, ValueError, smtplib.SMTPException) as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
